from flask import Flask, render_template, request, redirect, flash, url_for
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
app.secret_key = "lumens_ndp_secret_key"

EXCEL_FILE = "ndp_registrations.xlsx"

@app.route("/", methods=["GET", "POST"])
def registration():
    if request.method == "POST":
        name = request.form["customer_full_name"]
        contact_number = request.form["customer_contact_number"]
        carplate = request.form["vehicle_plate_number"]

        if not name or not contact_number or not carplate:
            flash("Please fill in all fields.", "error")
            return redirect(url_for("registration", _anchor="registration-form"))

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if os.path.exists(EXCEL_FILE):
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Registrations"
            sheet.append(["Timestamp", "Name", "Contact Number", "Car Plate Number"])

        sheet.append([timestamp, name, contact_number, carplate])
        workbook.save(EXCEL_FILE)

        flash("Registration submitted successfully.", "success")
        return redirect(url_for("registration", _anchor="registration-form"))

    return render_template("registration.html")

if __name__ == "__main__":
    app.run(debug=True)